# -*- coding: utf-8 -*-
"""
build_notes.py — توليد صفحات "مشاركات حلقة الفجر" من ملف Excel.

يقرأ fajr-notes.xlsx (في جذر المشروع)، يجمّع الصفوف حسب عمود "القسم"،
ويكتب صفحة Markdown لكل قسم داخل surahs/fajr/ مع ملف .pages للعنوان والترتيب.
الأقسام ديناميكية: أي قسم جديد تكتبه في Excel يصير صفحة جديدة تلقائيًا.

الأعمدة المتوقعة (الصف الأول رؤوس):
    القسم | التاريخ | العنوان | المحتوى | المرجع

التشغيل:
    python build_notes.py --init   # أنشئ قالب fajr-notes.xlsx (مرة واحدة)
    python build_notes.py          # ولّد الصفحات من الملف
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import Workbook, load_workbook

PROJECT_DIR = Path(__file__).resolve().parent.parent
XLSX = PROJECT_DIR / "fajr-notes.xlsx"
FAJR_DIR = PROJECT_DIR / "surahs" / "fajr"
SECTION_TITLE = "مشاركات حلقة الفجر"
HEADERS = ["القسم", "التاريخ", "العنوان", "المحتوى", "المرجع"]


def init_template() -> None:
    if XLSX.exists():
        print(f"الملف موجود بالفعل: {XLSX.name} — لن يُستبدل.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "notes"
    ws.append(HEADERS)
    ws.append(["فوائد لغوية", "2026-05-29", "الفرق بين الحِلم والأناة",
               "الحِلم: ضبط النفس عند الغضب. والأناة: التثبّت وترك العجلة.", "صحيح مسلم"])
    ws.append(["فوائد بلاغية", "2026-05-29", "تقديم المعمول للحصر",
               "في «إياك نعبد» قُدّم المفعول على الفعل لإفادة الحصر: لا نعبد إلا إياك.", "الفاتحة 5"])
    ws.append(["تجويد", "2026-05-29", "أحكام النون الساكنة",
               "الإظهار والإدغام والإقلاب والإخفاء، مع أمثلة من جزء عمّ.", ""])
    ws.append(["أسئلة وأجوبة", "2026-05-29", "متى بدأت دعوة عيسى عليه السلام؟",
               "خلاصة نقاش الحلقة حول بداية الدعوة...", ""])
    ws.append(["فوائد فقهية", "2026-05-29", "مسألة في الطهارة",
               "تفصيل ما اتفقنا عليه في الحلقة...", ""])
    ws.sheet_view.rightToLeft = True
    wb.save(XLSX)
    print(f"✓ أُنشئ القالب: {XLSX.name} — افتحه وأضف ملاحظاتك ثم شغّل: python build_notes.py")


def slug(name: str) -> str:
    return name.strip().replace("/", "-").replace(" ", "-")


def read_rows() -> list[dict]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: header.index(h) for h in HEADERS if h in header}

    def cell(row, col):
        i = idx.get(col)
        v = row[i] if i is not None and i < len(row) else None
        return str(v).strip() if v is not None else ""

    out = []
    for row in rows[1:]:
        category, content = cell(row, "القسم"), cell(row, "المحتوى")
        if not category or not content:
            continue
        out.append({
            "category": category,
            "date": cell(row, "التاريخ"),
            "title": cell(row, "العنوان"),
            "content": content,
            "ref": cell(row, "المرجع"),
        })
    return out


def build() -> None:
    if not XLSX.exists():
        print(f"✗ لا يوجد {XLSX.name}. أنشئه أولًا:  python build_notes.py --init")
        return
    rows = read_rows()
    if not rows:
        print("لا توجد ملاحظات في الملف بعد.")
        return

    FAJR_DIR.mkdir(parents=True, exist_ok=True)

    order: list[str] = []
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        c = row["category"]
        if c not in grouped:
            grouped[c] = []
            order.append(c)
        grouped[c].append(row)

    nav_files: list[str] = []
    for category in order:
        entries = sorted(grouped[category], key=lambda e: e["date"], reverse=True)
        lines = [f"# {category}\n"]
        for e in entries:
            meta = " — ".join(x for x in (e["date"], e["ref"]) if x)
            lines.append(f"## {e['title'] or '(بدون عنوان)'}")
            if meta:
                lines.append(f"*{meta}*")
            lines.append("")
            lines.append(e["content"])
            lines.append("\n---\n")
        fname = f"{slug(category)}.md"
        (FAJR_DIR / fname).write_text("\n".join(lines), encoding="utf-8")
        nav_files.append(fname)
        print(f"✓ {fname} — {len(entries)} ملاحظة")

    # إزالة صفحات أقسام لم تعد موجودة في Excel
    for stale in FAJR_DIR.glob("*.md"):
        if stale.name not in nav_files:
            stale.unlink()
            print(f"− حُذف قسم قديم: {stale.name}")

    pages = [f"title: {SECTION_TITLE}", "nav:"] + [f"  - {f}" for f in nav_files]
    (FAJR_DIR / ".pages").write_text("\n".join(pages) + "\n", encoding="utf-8")
    print(f"✓ .pages — {len(nav_files)} قسم")


def main() -> None:
    if "--init" in sys.argv:
        init_template()
    else:
        build()


if __name__ == "__main__":
    main()
